"""
Ledger Match — E-way Bill ⇄ Books Reconciliation
--------------------------------------------------
A Streamlit app that matches E-way Bill invoice values against Books,
voucher-wise, within a configurable tolerance.

Run with:
    pip install streamlit openpyxl pandas --break-system-packages
    streamlit run ledger_match_app.py
"""

import io
import re
import difflib
import hashlib
from copy import copy

import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

st.set_page_config(page_title="Ledger Match", page_icon="📒", layout="wide")

# ----------------------------------------------------------------------------
# Styling — ledger / register aesthetic
# ----------------------------------------------------------------------------
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&display=swap');
    html, body, [class*="css"] { font-family: -apple-system, "Segoe UI", sans-serif; }
    .ledger-title {
        font-family: Georgia, 'Times New Roman', serif;
        font-size: 40px;
        font-weight: 700;
        border-bottom: 3px double #16243B;
        padding-bottom: 10px;
        margin-bottom: 4px;
        color: #16243B;
    }
    .ledger-dek { color: #3A4A63; font-size: 14px; margin-bottom: 24px; }
    .section-label {
        font-family: 'IBM Plex Mono', monospace;
        font-size: 11px;
        letter-spacing: 2px;
        text-transform: uppercase;
        color: #9C7A2E;
        border-bottom: 1px solid #B9B096;
        padding-bottom: 6px;
        margin: 28px 0 12px 0;
    }
    .stamp-ok { color: #2E6B54; font-weight: 700; }
    .stamp-miss { color: #9E3A34; font-weight: 700; }
    div[data-testid="stMetricValue"] { font-family: 'IBM Plex Mono', monospace; }
    </style>
    <div class="ledger-title">Ledger Match</div>
    <div class="ledger-dek">Match E-way Bill invoice values against Books, voucher-wise — upload both sides, set a tolerance, and get a marked-up workbook back.</div>
    """,
    unsafe_allow_html=True,
)

# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------

def norm(s):
    """Normalize a header string for loose matching."""
    if s is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(s).strip().lower())


BOOKS_ALIASES = {
    "voucher_no": ["voucherno", "voucherno.", "vouchernumber", "voucher"],
    "invoice_value": ["invoicevalue", "invoiceamount", "invoiceamt"],
    "party_name": ["partyname", "party", "vendorname", "vendor"],
}
EWB_ALIASES = {
    "invoice_value": [
        "invoicevalue", "asperewaybillinvoicevalue", "asperewaybill",
        "ewbvalue", "ewayvillvalue", "invoiceamount",
    ],
    "ewb_no": ["ewbno"],
    "doc_no": ["docno"],
    "party_name": ["fromtradername", "partyname", "tradername", "fromparty"],
}


def parse_amount(v):
    """Parse a cell value into a float, handling commas/currency/brackets. None if not parseable."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    s = re.sub(r"[₹$,\s]", "", s)
    try:
        n = float(s)
    except ValueError:
        return None
    return -n if neg else n


def cell_text(v):
    if v is None:
        return ""
    return str(v)


_CORP_SUFFIXES = re.compile(
    r"\b(PRIVATE|PVT|LIMITED|LTD|LLP|INC|CO|COMPANY|ENGG|ENGINEERING|"
    r"ENTERPRISES|WORKS|INDUSTRIES|TRADERS|CORP|CORPORATION)\b"
)


def canonical_key(s):
    """Light-touch normalization used to bucket rows into distinct parties WITHIN one file:
    case, punctuation and whitespace differences collapse together, but real words are kept
    intact. Deliberately does NOT strip corporate-suffix words like normalize_name() does —
    doing so here would wrongly merge genuinely different companies that happen to share a
    root name with different suffixes (e.g. "Sri Balaji Traders" vs "Sri Balaji Enterprises")."""
    if not s:
        return ""
    s = str(s).upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_name(s):
    """Aggressive normalization — uppercase, strip punctuation AND common corporate suffixes —
    used only for scoring SIMILARITY between a Books name and an E-way Bill name across the two
    files (e.g. recognizing "Ganapathi Engg Works" and "Ganapathi Engineering Works Pvt Ltd" as
    likely the same company). Never use this as a dict key for bucketing distinct parties within
    a single file — that's what canonical_key() is for."""
    if not s:
        return ""
    s = str(s).upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = _CORP_SUFFIXES.sub(" ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def name_similarity(a, b):
    """0-1 similarity score between two party names, robust to minor spelling/formatting differences."""
    na, nb = normalize_name(a), normalize_name(b)
    if not na or not nb:
        return 0.0
    return difflib.SequenceMatcher(None, na, nb).ratio()


def pick_main_sheet(wb):
    """Pick the worksheet with the most rows."""
    return max(wb.worksheets, key=lambda ws: ws.max_row)


def detect_columns(ws, alias_map):
    """Return {key: column_index} for headers found in row 1, matched against alias_map."""
    found = {}
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    for cell in header_row:
        h = norm(cell.value)
        if not h:
            continue
        for key, aliases in alias_map.items():
            if key in found:
                continue
            if h in aliases:
                found[key] = cell.column
    return found


def get_or_create_col(ws, header_text, norm_key):
    """Find an existing column with this normalized header, or append a new one."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    for cell in header_row:
        if norm(cell.value) == norm_key:
            return cell.column
    col = ws.max_column + 1
    hc = ws.cell(row=1, column=col, value=header_text)
    hc.font = Font(bold=True, color="FFFFFFFF")
    hc.fill = PatternFill(start_color="FF16243B", end_color="FF16243B", fill_type="solid")
    return col


def load_workbook_from_upload(uploaded_file):
    # getvalue() is idempotent (unlike read(), which consumes the stream) — safe to call
    # again on later reruns without the file appearing empty.
    data = uploaded_file.getvalue()
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = pick_main_sheet(wb)
    return wb, ws


def build_template(kind: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    if kind == "books":
        ws.title = "Books"
        headers = ["Voucher No", "Date", "Party Name", "Ledger", "Invoice Value", "Narration"]
        sample = ["19004", "2026-04-05", "Example Vendor Pvt Ltd", "Purchase Imports",
                  "8,60,366.44", "Sample row — replace with your data"]
    else:
        ws.title = "E way bills"
        headers = ["EWB No", "Doc No", "Doc Date", "Supply Type", "From Trader Name", "Invoice Value"]
        sample = ["4123456789012", "XIVBLR300310161", "2026-04-05", "Outward",
                  "Example Buyer Pvt Ltd", "8,60,366.44"]
    ws.append(headers)
    ws.append(sample)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(start_color="FF16243B", end_color="FF16243B", fill_type="solid")
    for cell in ws[2]:
        cell.font = Font(italic=True, color="FF808080")
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_groups_and_ewb_rows(books_ws, books_cols, ewb_ws, ewb_cols, use_party):
    """Read both sheets into memory: Books voucher groups and E-way Bill rows.

    Also does a full, independent scan of every row's Party Name — separate from the
    voucher-grouping logic below — so a party is never silently dropped from the reference
    list just because its row had a blank Voucher No / unparseable Invoice Value, or because
    it shared a Voucher No with a different party (which only the *first* row's party gets
    credited with, for summing purposes).
    """
    all_books_parties = {}   # canonical_key -> first-seen original text
    all_ewb_parties = {}

    groups = {}
    for row in books_ws.iter_rows(min_row=2):
        if use_party and "party_name" in books_cols:
            pname_raw = cell_text(row[books_cols["party_name"] - 1].value).strip()
            if pname_raw:
                ck = canonical_key(pname_raw)
                if ck and ck not in all_books_parties:
                    all_books_parties[ck] = pname_raw

        v_cell = row[books_cols["voucher_no"] - 1]
        vtext = cell_text(v_cell.value).strip()
        if not vtext:
            continue
        amt = parse_amount(row[books_cols["invoice_value"] - 1].value)
        if amt is None:
            continue
        g = groups.setdefault(vtext, {"rows": [], "sum": 0.0, "party": ""})
        g["rows"].append(v_cell.row)
        g["sum"] += amt
        if not g["party"] and "party_name" in books_cols:
            pname = cell_text(row[books_cols["party_name"] - 1].value).strip()
            if pname:
                g["party"] = pname
    for g in groups.values():
        g["sum"] = round(g["sum"], 2)

    ewb_rows = []
    for row in ewb_ws.iter_rows(min_row=2):
        if use_party and "party_name" in ewb_cols:
            pname_raw = cell_text(row[ewb_cols["party_name"] - 1].value).strip()
            if pname_raw:
                ck = canonical_key(pname_raw)
                if ck and ck not in all_ewb_parties:
                    all_ewb_parties[ck] = pname_raw

        cell = row[ewb_cols["invoice_value"] - 1]
        amt = parse_amount(cell.value)
        if amt is None:
            continue
        ref = ""
        if "ewb_no" in ewb_cols:
            ref = cell_text(row[ewb_cols["ewb_no"] - 1].value).strip()
        if not ref and "doc_no" in ewb_cols:
            ref = cell_text(row[ewb_cols["doc_no"] - 1].value).strip()
        if not ref:
            ref = f"Row {cell.row}"
        doc_no = cell_text(row[ewb_cols["doc_no"] - 1].value).strip() if "doc_no" in ewb_cols else ""
        party = ""
        if "party_name" in ewb_cols:
            party = cell_text(row[ewb_cols["party_name"] - 1].value).strip()
        ewb_rows.append({"r": cell.row, "val": round(amt, 2), "ref": ref, "party": party, "doc_no": doc_no})

    for vn, g in groups.items():
        g["bkey"] = canonical_key(g["party"]) if use_party else ""
        g["vn_key"] = canonical_key(vn)
    for er in ewb_rows:
        er["ekey"] = canonical_key(er["party"]) if use_party else ""
        er["doc_key"] = canonical_key(er["doc_no"])

    return groups, ewb_rows, all_books_parties, all_ewb_parties


def stage1_party_pairing(groups, ewb_rows, all_books_parties, all_ewb_parties, party_threshold):
    """Segregate Books/E-way Bills by party and auto-suggest pairings above `party_threshold`.

    book_reps/ewb_reps are the FULL set of distinct party names found anywhere in each file
    (from `all_books_parties`/`all_ewb_parties`) — not just the ones that made it into a valid,
    summable voucher group. A party can therefore appear here with 0 vouchers/invoices if all of
    its rows were skipped (blank Voucher No, unparseable Invoice Value) or if it shared a Voucher
    No with a different party that got credited instead — which is worth surfacing, not hiding.

    Returns everything needed both to render an editable review table and, later, to run
    Stage 2 with whatever mapping the person confirms.
    """
    book_reps = dict(all_books_parties)
    ewb_reps = dict(all_ewb_parties)
    vouchers_by_bkey = {}
    ewb_rows_by_ekey = {}

    for vn, g in groups.items():
        k = g["bkey"]
        if not k:
            continue
        vouchers_by_bkey.setdefault(k, []).append(vn)

    for er in ewb_rows:
        k = er["ekey"]
        if not k:
            continue
        ewb_rows_by_ekey.setdefault(k, []).append(er)

    candidates = []
    for bk, bname in book_reps.items():
        for ek, ename in ewb_reps.items():
            sim = name_similarity(bname, ename)
            if sim * 100 >= party_threshold:
                candidates.append((sim, bk, ek))
    candidates.sort(key=lambda c: -c[0])

    used_b, used_e = set(), set()
    auto_ekey_to_bkey, sim_by_ekey = {}, {}
    for sim, bk, ek in candidates:
        if bk in used_b or ek in used_e:
            continue
        auto_ekey_to_bkey[ek] = bk
        sim_by_ekey[ek] = sim
        used_b.add(bk)
        used_e.add(ek)

    return {
        "book_reps": book_reps,
        "ewb_reps": ewb_reps,
        "vouchers_by_bkey": vouchers_by_bkey,
        "ewb_rows_by_ekey": ewb_rows_by_ekey,
        "auto_ekey_to_bkey": auto_ekey_to_bkey,
        "sim_by_ekey": sim_by_ekey,
    }


def stage2_amount_match(groups, ewb_rows, use_party, ekey_to_bkeys, vouchers_by_bkey, tolerance,
                         use_voucher_no=False, use_amount=True, book_reps=None):
    """Greedy-match Books voucher groups against E-way Bill rows using whichever criteria are
    enabled. Each voucher is used at most once.

    - use_amount: require the Books voucher sum to be within `tolerance` of the E-way Bill
      Invoice Value. Amount closeness is always used to break ties between otherwise-equal
      candidates, even when this is off.
    - use_voucher_no: require the Books Voucher No. to (loosely) equal the E-way Bill's Doc No.
    - use_party: require the E-way Bill's party to be paired (via ekey_to_bkeys) to the Books
      voucher's party. One E-way Bill party can map to several Books party keys (e.g. multiple
      branches of the same vendor).

    At least one of the three should be True or nothing will ever be filtered — this is enforced
    by the caller (the UI won't let all three be switched off).
    """
    book_reps = book_reps or {}
    all_vns = list(groups.keys())

    def vouchers_for_bkeys(bkeys):
        vns = []
        for bk in bkeys:
            vns.extend(vouchers_by_bkey.get(bk, []))
        return vns

    def candidate_pool(er):
        if use_party:
            pool = vouchers_for_bkeys(ekey_to_bkeys.get(er["ekey"], []))
        else:
            pool = all_vns
        if use_voucher_no:
            dk = er["doc_key"]
            pool = [vn for vn in pool if dk and groups[vn]["vn_key"] == dk]
        return pool

    pairs = []
    for er in ewb_rows:
        for vn in candidate_pool(er):
            g = groups[vn]
            d = abs(g["sum"] - er["val"])
            if use_amount and d > tolerance:
                continue
            pairs.append({"d": d, "r": er["r"], "val": er["val"], "vn": vn,
                          "sum": g["sum"], "ref": er["ref"],
                          "books_party": g["party"], "ewb_party": er["party"]})
    pairs.sort(key=lambda p: p["d"])

    used_vn, used_row = set(), set()
    matches = []
    for p in pairs:
        if p["vn"] in used_vn or p["r"] in used_row:
            continue
        matches.append(p)
        used_vn.add(p["vn"])
        used_row.add(p["r"])

    matched_rows = {m["r"] for m in matches}
    unmatched_ewb = [er for er in ewb_rows if er["r"] not in matched_rows]

    unmatched_diag = []
    for er in unmatched_ewb:
        pool = candidate_pool(er)
        status_bits = []
        if use_party:
            if not er["ekey"]:
                status_bits.append("no party name on this row")
            else:
                bkeys = ekey_to_bkeys.get(er["ekey"], [])
                if not bkeys:
                    status_bits.append("no matching Books party confirmed")
                else:
                    names = ", ".join(f"\u201c{book_reps.get(bk, bk)}\u201d" for bk in bkeys)
                    status_bits.append(f"party matched to Books {names}")
        if use_voucher_no:
            if not er["doc_key"]:
                status_bits.append("no Doc No on this row")
            elif not pool:
                status_bits.append("no Books Voucher No. matches this Doc No.")
        best = None
        for vn in pool:
            d = abs(groups[vn]["sum"] - er["val"])
            if best is None or d < best["d"]:
                best = {"d": d, "vn": vn, "sum": groups[vn]["sum"]}
        if best is None and not pool and not status_bits:
            status_bits.append("no candidate rows found")
        unmatched_diag.append({**er, "best": best,
                               "party_status": "; ".join(status_bits) if status_bits else None})

    unused_vouchers = [{"vn": vn, "sum": g["sum"], "party": g["party"]}
                       for vn, g in groups.items() if vn not in used_vn]

    used_bkeys = {bk for bkeys in ekey_to_bkeys.values() for bk in bkeys} if use_party else set()
    unmatched_books_parties = [
        {"party": bname, "vouchers": len(vouchers_by_bkey.get(bk, [])),
         "sum": round(sum(groups[vn]["sum"] for vn in vouchers_by_bkey.get(bk, [])), 2)}
        for bk, bname in book_reps.items() if bk not in used_bkeys
    ]

    return {
        "groups": groups,
        "matches": matches,
        "unmatched_diag": unmatched_diag,
        "unused_vouchers": unused_vouchers,
        "total_ewb": len(ewb_rows),
        "use_party": use_party,
        "unmatched_books_parties": unmatched_books_parties,
    }


def workbook_to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def fmt(n):
    return f"{n:,.2f}"


MAPPING_ALIASES = {
    "ewb_party": ["ewaybillparty", "ewbparty", "asperewaybillparty", "ewaybillpartyname"],
    "books_party": ["booksparty", "asperbooksparty", "bookspartyname"],
}


def build_party_mapping_template_bytes(s1) -> bytes:
    """A 4-sheet workbook: instructions, the editable mapping (pre-filled with auto-suggestions),
    and two reference sheets listing every distinct party name found in each file — so the person
    can copy exact spellings instead of retyping them."""
    wb = openpyxl.Workbook()

    ws0 = wb.active
    ws0.title = "Instructions"
    for line in [
        "How to fill this in",
        "",
        "1. Open the 'Party Mapping' sheet. Each row already has an E-way Bill Party name.",
        "2. Fill in (or correct) the matching Books Party name in column B.",
        "3. Copy exact spellings from 'Books Parties (reference)' / 'E-way Bill Parties (reference)' "
        "to avoid typos — the app matches on exact (normalized) name.",
        "4. To map ONE E-way Bill Party to MULTIPLE Books Party names (e.g. the same vendor recorded "
        "under different branch names), add another row: repeat the same E-way Bill Party in column A, "
        "with a different Books Party in column B.",
        "5. Leave column B blank for any E-way Bill Party that has no match in Books.",
        "6. Save this file and upload it back into the app.",
    ]:
        ws0.append([line])
    ws0.column_dimensions["A"].width = 110
    ws0["A1"].font = Font(bold=True, size=13)

    ewb_keys_sorted = sorted(s1["ewb_reps"].keys(), key=lambda ek: s1["ewb_reps"][ek])

    ws = wb.create_sheet("Party Mapping")
    ws.append(["E-way Bill Party", "Books Party", "Invoices", "Total (E-way Bill)"])
    for ek in ewb_keys_sorted:
        rows_for = s1["ewb_rows_by_ekey"].get(ek, [])
        auto_bk = s1["auto_ekey_to_bkey"].get(ek)
        books_val = s1["book_reps"][auto_bk] if auto_bk else ""
        ws.append([s1["ewb_reps"][ek], books_val, len(rows_for),
                   round(sum(r["val"] for r in rows_for), 2)])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(start_color="FF16243B", end_color="FF16243B", fill_type="solid")
    for col, width in zip("ABCD", (42, 42, 11, 18)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Books Parties (reference)")
    ws2.append(["Books Party", "Vouchers"])
    for bk, bname in sorted(s1["book_reps"].items(), key=lambda kv: kv[1]):
        ws2.append([bname, len(s1["vouchers_by_bkey"].get(bk, []))])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(start_color="FF16243B", end_color="FF16243B", fill_type="solid")
    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 12
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("E-way Bill Parties (reference)")
    ws3.append(["E-way Bill Party", "Invoices", "Total"])
    for ek in ewb_keys_sorted:
        rows_for = s1["ewb_rows_by_ekey"].get(ek, [])
        ws3.append([s1["ewb_reps"][ek], len(rows_for), round(sum(r["val"] for r in rows_for), 2)])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(start_color="FF16243B", end_color="FF16243B", fill_type="solid")
    ws3.column_dimensions["A"].width = 45
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 16
    ws3.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def find_mapping_sheet(wb):
    """Prefer a sheet literally named 'Party Mapping'; otherwise the first sheet whose header
    row matches the mapping column aliases; otherwise fall back to the largest sheet."""
    for ws in wb.worksheets:
        if norm(ws.title) == norm("Party Mapping"):
            return ws
    for ws in wb.worksheets:
        if detect_columns(ws, MAPPING_ALIASES).keys() >= {"ewb_party", "books_party"}:
            return ws
    return pick_main_sheet(wb)


def _resolve_typed_name(text, reps, name_to_key):
    """Resolve a typed party name back to its key: exact canonical match first; if that fails,
    fall back to the single best fuzzy match, but only if it's confident (>=90% similarity) —
    tolerates minor typos without risking an ambiguous/wrong merge."""
    key = name_to_key.get(canonical_key(text))
    if key is not None:
        return key, False
    best_key, best_sim = None, 0.0
    for k, name in reps.items():
        sim = name_similarity(text, name)
        if sim > best_sim:
            best_key, best_sim = k, sim
    if best_key is not None and best_sim >= 0.90:
        return best_key, True
    return None, False


def parse_party_mapping_upload(ws, ewb_reps, book_reps):
    """Read the person's completed mapping file and resolve each row to the actual E-way Bill /
    Books party keys from the currently-loaded files. Returns (ekey_to_bkeys, warnings) —
    warnings list any row that couldn't be resolved, e.g. because the typed name doesn't (even
    loosely) match anything in the uploaded Books/E-way Bill files.
    """
    cols = detect_columns(ws, MAPPING_ALIASES)
    if "ewb_party" not in cols or "books_party" not in cols:
        return None, ["Couldn't find 'E-way Bill Party' and 'Books Party' column headers in row 1 "
                       "— use the downloaded template without renaming the header row."]

    ewb_name_to_key = {canonical_key(name): ek for ek, name in ewb_reps.items()}
    book_name_to_key = {canonical_key(name): bk for bk, name in book_reps.items()}

    ekey_to_bkeys = {}
    warnings = []
    for row in ws.iter_rows(min_row=2):
        ewb_text = cell_text(row[cols["ewb_party"] - 1].value).strip()
        books_text = cell_text(row[cols["books_party"] - 1].value).strip()
        if not ewb_text:
            continue
        ek, fuzzy = _resolve_typed_name(ewb_text, ewb_reps, ewb_name_to_key)
        if ek is None:
            warnings.append(f"Row {row[0].row}: E-way Bill party \u201c{ewb_text}\u201d not found in "
                             "the uploaded E-way Bill file — skipped.")
            continue
        if fuzzy:
            warnings.append(f"Row {row[0].row}: E-way Bill party \u201c{ewb_text}\u201d matched "
                             f"approximately to \u201c{ewb_reps[ek]}\u201d — check this is correct.")
        if not books_text:
            continue  # explicitly left blank = no match, and that's fine
        bk, fuzzy = _resolve_typed_name(books_text, book_reps, book_name_to_key)
        if bk is None:
            warnings.append(f"Row {row[0].row}: Books party \u201c{books_text}\u201d not found in "
                             "the uploaded Books file — skipped.")
            continue
        if fuzzy:
            warnings.append(f"Row {row[0].row}: Books party \u201c{books_text}\u201d matched "
                             f"approximately to \u201c{book_reps[bk]}\u201d — check this is correct.")
        bucket = ekey_to_bkeys.setdefault(ek, [])
        if bk not in bucket:
            bucket.append(bk)

    return ekey_to_bkeys, warnings


# ----------------------------------------------------------------------------
# 1. Templates
# ----------------------------------------------------------------------------
st.markdown('<div class="section-label">1 · Get the format right</div>', unsafe_allow_html=True)
c1, c2 = st.columns(2)
with c1:
    st.markdown("**Books template**")
    st.caption("One row per ledger line. Rows sharing a Voucher No. are summed together.")
    st.markdown(
        "- **Voucher No** — required\n"
        "- **Invoice Value** — required\n"
        "- Date, **Party Name**, Ledger, Narration — optional (Party Name needed for approx. name matching)"
    )
    st.download_button(
        "Download Books template",
        data=build_template("books"),
        file_name="Books - Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
with c2:
    st.markdown("**E-way Bill template**")
    st.caption("One row per E-way bill. Its Invoice Value is what gets matched against Books.")
    st.markdown(
        "- **Invoice Value** — required\n"
        "- EWB No, Doc No, Doc Date, Supply Type, **From Trader Name** — optional (needed for approx. name matching)"
    )
    st.download_button(
        "Download E-way Bill template",
        data=build_template("ewb"),
        file_name="E way Bill - Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

def file_signature(uploaded_file):
    """A content-based fingerprint for an uploaded file (hash of its actual bytes), used to
    detect when the person has swapped in a genuinely different file — independent of any
    Streamlit-internal file ID quirks, and independent of filename/size coincidences."""
    return hashlib.md5(uploaded_file.getvalue()).hexdigest()


def clear_downstream_state(clear_workbooks=False):
    """Wipe everything derived from the previous Books/E-way Bill files: party pairing,
    the review-table widgets, and any final results — so a fresh upload never shows stale
    data from a prior file."""
    keys = ["stage1", "result", "groups", "ewb_rows", "use_party_used", "use_voucher_no_used",
            "use_amount_used", "tolerance_used", "mapping_upload"]
    if clear_workbooks:
        keys += ["books_wb", "books_ws_title", "books_cols",
                 "ewb_wb", "ewb_ws_title", "ewb_cols",
                 "books_file_sig", "ewb_file_sig"]
    for k in keys:
        st.session_state.pop(k, None)
    st.session_state["mapping_version"] = st.session_state.get("mapping_version", 0) + 1


# ----------------------------------------------------------------------------
# 2. Uploads
# ----------------------------------------------------------------------------
st.markdown('<div class="section-label">2 · Upload both sides</div>', unsafe_allow_html=True)

_, reset_col = st.columns([5, 1])
with reset_col:
    if st.button("🔄 Reset everything"):
        clear_downstream_state(clear_workbooks=True)
        st.rerun()

u1, u2 = st.columns(2)
with u1:
    books_file = st.file_uploader("Books (.xlsx)", type=["xlsx"], key="books_upload")
with u2:
    ewb_file = st.file_uploader("E-way Bills (.xlsx)", type=["xlsx"], key="ewb_upload")

books_ready = ewb_ready = False
books_cols = ewb_cols = None

if books_file is not None:
    sig = file_signature(books_file)
    if st.session_state.get("books_file_sig") != sig:
        clear_downstream_state()
        st.session_state["books_file_sig"] = sig

    books_wb, books_ws = load_workbook_from_upload(books_file)
    books_cols = detect_columns(books_ws, BOOKS_ALIASES)
    missing = [k for k in ("voucher_no", "invoice_value") if k not in books_cols]
    if missing:
        st.error(
            f"Books file: couldn't find {', '.join(missing)} column header in row 1. "
            "Use the Books template headers."
        )
    else:
        st.session_state["books_wb"] = books_wb
        st.session_state["books_ws_title"] = books_ws.title
        st.session_state["books_cols"] = books_cols
        books_ready = True
        party_note = " (Party Name detected)" if "party_name" in books_cols else " (no Party Name column found)"
        st.success(f"Books: {books_ws.max_row - 1} rows detected{party_note}")

if ewb_file is not None:
    sig = file_signature(ewb_file)
    if st.session_state.get("ewb_file_sig") != sig:
        clear_downstream_state()
        st.session_state["ewb_file_sig"] = sig

    ewb_wb, ewb_ws = load_workbook_from_upload(ewb_file)
    ewb_cols = detect_columns(ewb_ws, EWB_ALIASES)
    if "invoice_value" not in ewb_cols:
        st.error(
            "E-way Bill file: couldn't find an 'Invoice Value' column header in row 1. "
            "Use the E-way Bill template headers."
        )
    else:
        st.session_state["ewb_wb"] = ewb_wb
        st.session_state["ewb_ws_title"] = ewb_ws.title
        st.session_state["ewb_cols"] = ewb_cols
        ewb_ready = True
        party_note = " (Party Name detected)" if "party_name" in ewb_cols else " (no Party Name column found)"
        st.success(f"E-way Bills: {ewb_ws.max_row - 1} rows detected{party_note}")

# ----------------------------------------------------------------------------
# 3. Matching rules & run
# ----------------------------------------------------------------------------
# ----------------------------------------------------------------------------
# 3. Matching rules
# ----------------------------------------------------------------------------
st.markdown('<div class="section-label">3 · Set matching criteria</div>', unsafe_allow_html=True)

party_cols_ok = bool(books_cols and "party_name" in books_cols and ewb_cols and "party_name" in ewb_cols)
voucher_cols_ok = bool(ewb_cols and "doc_no" in ewb_cols)

st.write("Match on:")
cc1, cc2, cc3 = st.columns(3)
with cc1:
    crit_amount = st.checkbox(
        "Amount", value=True,
        help="Books voucher sum must be within your chosen tolerance of the E-way Bill Invoice Value.",
    )
with cc2:
    crit_voucher = st.checkbox(
        "Voucher No.", value=False, disabled=not voucher_cols_ok,
        help="Books Voucher No. must match the E-way Bill's Doc No. (case/punctuation-insensitive). "
             "Needs a Doc No column in the E-way Bill file.",
    )
with cc3:
    crit_party = st.checkbox(
        "Party Name", value=party_cols_ok, disabled=not party_cols_ok,
        help="Groups Books by Party Name and E-way Bills by From Trader Name, and only matches "
             "amounts within a confirmed party pairing. Prevents unrelated parties' invoices from "
             "matching just because the amounts happen to be close.",
    )

if not (crit_amount or crit_voucher or crit_party):
    st.error("Select at least one matching criterion.")
if not voucher_cols_ok:
    st.caption("To enable Voucher No. matching, include a **Doc No** column in the E-way Bill file.")
if not party_cols_ok:
    st.caption(
        "To enable Party Name matching, include a **Party Name** column in Books and a "
        "**From Trader Name** column in E-way Bills."
    )

use_amount, use_voucher_no, use_party = crit_amount, crit_voucher, crit_party

party_threshold = 55.0
if use_party:
    party_threshold = st.slider(
        "Minimum name similarity to auto-suggest a pairing",
        min_value=0, max_value=100, value=55, step=5,
        help="Party names are compared after stripping punctuation and common suffixes "
             "(Pvt, Ltd, Enterprises, Engg, Works, etc.). This only affects the starting suggestions — "
             "you can override any of them by hand in the next step, either way.",
    )

tolerance = 0.0
if use_amount:
    t1, t2 = st.columns([2, 1])
    with t1:
        preset = st.radio(
            "Treat amounts as matched when within",
            ["Exact (₹0)", "₹1", "₹100", "₹500", "₹1,000", "Custom"],
            index=1,
            horizontal=True,
        )
    with t2:
        custom_tol = st.number_input("Custom ₹ tolerance", min_value=0.0, value=0.0, step=1.0,
                                      disabled=(preset != "Custom"))
    preset_map = {"Exact (₹0)": 0.0, "₹1": 1.0, "₹100": 100.0, "₹500": 500.0, "₹1,000": 1000.0}
    tolerance = custom_tol if preset == "Custom" else preset_map[preset]
else:
    st.caption(
        "Amount isn't a selected criterion — the closest-amount candidate is still used to break "
        "ties when several Books vouchers otherwise match equally well, but nothing is required "
        "to be within any particular ₹ range."
    )

find_clicked = st.button(
    "Prepare party mapping →" if use_party else "Run reconciliation →",
    type="primary", disabled=not (books_ready and ewb_ready and (crit_amount or crit_voucher or crit_party)),
)

if find_clicked:
    groups, ewb_rows, all_books_parties, all_ewb_parties = build_groups_and_ewb_rows(
        st.session_state["books_wb"].worksheets[
            [ws.title for ws in st.session_state["books_wb"].worksheets].index(st.session_state["books_ws_title"])
        ],
        st.session_state["books_cols"],
        st.session_state["ewb_wb"].worksheets[
            [ws.title for ws in st.session_state["ewb_wb"].worksheets].index(st.session_state["ewb_ws_title"])
        ],
        st.session_state["ewb_cols"],
        use_party,
    )
    st.session_state["groups"] = groups
    st.session_state["ewb_rows"] = ewb_rows
    st.session_state["tolerance_used"] = tolerance
    st.session_state["use_party_used"] = use_party
    st.session_state["use_voucher_no_used"] = use_voucher_no
    st.session_state["use_amount_used"] = use_amount
    st.session_state.pop("result", None)

    if use_party:
        st.session_state["stage1"] = stage1_party_pairing(
            groups, ewb_rows, all_books_parties, all_ewb_parties, party_threshold
        )
        st.session_state["mapping_version"] = st.session_state.get("mapping_version", 0) + 1
    else:
        # No party segregation — go straight to amount/voucher-no matching across everything.
        result = stage2_amount_match(groups, ewb_rows, use_party=False, ekey_to_bkeys={},
                                      vouchers_by_bkey={}, tolerance=tolerance,
                                      use_voucher_no=use_voucher_no, use_amount=use_amount)
        st.session_state["result"] = result

# ----------------------------------------------------------------------------
# 4. Party matching — manual, via a downloadable/uploadable mapping file
# ----------------------------------------------------------------------------
if "stage1" in st.session_state and st.session_state.get("use_party_used"):
    st.markdown('<div class="section-label">4 · Party matching</div>', unsafe_allow_html=True)
    s1 = st.session_state["stage1"]

    st.caption(
        f"Found **{len(s1['ewb_reps'])}** distinct E-way Bill parties and **{len(s1['book_reps'])}** "
        "distinct Books parties. Download the template below (pre-filled with auto-suggested matches "
        "where found), correct or complete it in Excel — where you can see and copy exact spellings — "
        "then upload it back here."
    )

    zero_books = [name for bk, name in s1["book_reps"].items() if not s1["vouchers_by_bkey"].get(bk)]
    zero_ewb = [name for ek, name in s1["ewb_reps"].items() if not s1["ewb_rows_by_ekey"].get(ek)]
    if zero_books or zero_ewb:
        with st.expander(
            f"⚠️ {len(zero_books) + len(zero_ewb)} party name(s) found with no usable rows — click for details",
        ):
            st.caption(
                "These party names appear in the Party Name column but every row they're on was "
                "skipped — usually a blank Voucher No / Invoice Value, or the row shares a Voucher No "
                "with a different party whose name got credited with the total instead. They'll still "
                "show up for mapping below, but with 0 invoices/vouchers, so nothing will actually "
                "match under them. Worth checking the source rows."
            )
            if zero_books:
                st.write("**Books:**", ", ".join(zero_books))
            if zero_ewb:
                st.write("**E-way Bills:**", ", ".join(zero_ewb))

    dl_col, up_col = st.columns(2)
    with dl_col:
        st.download_button(
            "⬇ Download party mapping template",
            data=build_party_mapping_template_bytes(s1),
            file_name="Party Mapping - Template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with up_col:
        mapping_file = st.file_uploader("Upload completed party mapping (.xlsx)", type=["xlsx"],
                                         key="mapping_upload")

    if mapping_file is not None:
        mapping_wb = openpyxl.load_workbook(io.BytesIO(mapping_file.getvalue()))
        mapping_ws = find_mapping_sheet(mapping_wb)
        final_ekey_to_bkeys, warnings = parse_party_mapping_upload(mapping_ws, s1["ewb_reps"], s1["book_reps"])

        if final_ekey_to_bkeys is None:
            st.error(warnings[0])
        else:
            mapped_count = len(final_ekey_to_bkeys)
            st.success(
                f"Parsed the mapping: {mapped_count} of {len(s1['ewb_reps'])} E-way Bill parties "
                "have at least one Books party assigned."
            )
            if warnings:
                with st.expander(f"⚠️ {len(warnings)} row(s) couldn't be resolved — click for details"):
                    for w in warnings:
                        st.write("- " + w)

            with st.expander("Preview parsed mapping"):
                st.dataframe(
                    [
                        {"E-way Bill Party": s1["ewb_reps"][ek],
                         "Books Part(y/ies)": ", ".join(s1["book_reps"][bk] for bk in bks)}
                        for ek, bks in sorted(final_ekey_to_bkeys.items(), key=lambda kv: s1["ewb_reps"][kv[0]])
                    ],
                    use_container_width=True, hide_index=True,
                )

            if st.button("Confirm mapping & run amount matching →", type="primary"):
                result = stage2_amount_match(
                    st.session_state["groups"], st.session_state["ewb_rows"],
                    use_party=True, ekey_to_bkeys=final_ekey_to_bkeys,
                    vouchers_by_bkey=s1["vouchers_by_bkey"], tolerance=st.session_state["tolerance_used"],
                    use_voucher_no=st.session_state.get("use_voucher_no_used", False),
                    use_amount=st.session_state.get("use_amount_used", True),
                    book_reps=s1["book_reps"],
                )
                result["party_pairs"] = [
                    {"books_party": s1["book_reps"][bk], "ewb_party": s1["ewb_reps"][ek]}
                    for ek, bks in final_ekey_to_bkeys.items() for bk in bks
                ]
                st.session_state["result"] = result

# ----------------------------------------------------------------------------
# 5. Results
# ----------------------------------------------------------------------------
if "result" in st.session_state:
    res = st.session_state["result"]
    st.markdown('<div class="section-label">5 · Results</div>', unsafe_allow_html=True)

    if res["use_party"]:
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Parties confirmed", len(res["party_pairs"]))
        m2.metric("Matched invoices", f"{len(res['matches'])} / {res['total_ewb']}")
        m3.metric("Unmatched E-way Bills", len(res["unmatched_diag"]))
        m4.metric("Unused Books vouchers", len(res["unused_vouchers"]))

        with st.expander(f"Confirmed party mapping ({len(res['party_pairs'])})"):
            st.dataframe(
                [{"Books party": p["books_party"], "E-way Bill party": p["ewb_party"]}
                 for p in sorted(res["party_pairs"], key=lambda p: p["books_party"])],
                use_container_width=True, hide_index=True,
            )
            if res["unmatched_books_parties"]:
                st.caption("Books parties left with no E-way Bill counterpart:")
                st.dataframe(
                    [{"Party": u["party"], "Vouchers": u["vouchers"], "Total": fmt(u["sum"])}
                     for u in res["unmatched_books_parties"]],
                    use_container_width=True, hide_index=True,
                )
    else:
        m1, m2, m3 = st.columns(3)
        m1.metric("Matched", f"{len(res['matches'])} / {res['total_ewb']}")
        m2.metric("Unmatched E-way Bills", len(res["unmatched_diag"]))
        m3.metric("Unused Books vouchers", len(res["unused_vouchers"]))

    with st.expander(f"Matched invoices ({len(res['matches'])})", expanded=True):
        rows = sorted(res["matches"], key=lambda m: m["r"])
        table_rows = []
        for m in rows:
            row = {
                "Voucher No": m["vn"],
                "EWB No": m["ref"],
                "Books sum": fmt(m["sum"]),
                "E-way Bill value": fmt(m["val"]),
                "Diff": fmt(abs(m["sum"] - m["val"])),
            }
            if res["use_party"]:
                row["Party"] = m["books_party"]
            table_rows.append(row)
        st.dataframe(table_rows, use_container_width=True, hide_index=True)

    with st.expander(f"Unmatched E-way Bills ({len(res['unmatched_diag'])})"):
        table_rows = []
        for u in res["unmatched_diag"]:
            row = {
                "Reference": u["ref"],
                "Invoice Value": fmt(u["val"]),
            }
            if res["use_party"]:
                row["Party"] = u["party"]
            if u["party_status"]:
                row["Status"] = u["party_status"]
            row["Closest book voucher"] = u["best"]["vn"] if u["best"] else "—"
            row["Nearest diff"] = fmt(u["best"]["d"]) if u["best"] else "—"
            table_rows.append(row)
        st.dataframe(table_rows, use_container_width=True, hide_index=True)

    with st.expander(f"Unused Books vouchers ({len(res['unused_vouchers'])})"):
        table_rows = [
            {"Voucher No": u["vn"], "Books sum": fmt(u["sum"]),
             **({"Party": u["party"]} if res["use_party"] else {})}
            for u in res["unused_vouchers"]
        ]
        st.dataframe(table_rows, use_container_width=True, hide_index=True)

    # ---- Downloads ----
    d1, d2 = st.columns(2)

    with d1:
        books_wb = st.session_state["books_wb"]
        books_ws = books_wb.worksheets[
            [ws.title for ws in books_wb.worksheets].index(st.session_state["books_ws_title"])
        ]
        ewb_col = get_or_create_col(books_ws, "EWB No", "ewbno")
        # Reset the column first so a rerun at a different tolerance never leaves stale values
        for r in range(2, books_ws.max_row + 1):
            books_ws.cell(row=r, column=ewb_col).value = None
        for m in res["matches"]:
            for r in res["groups"][m["vn"]]["rows"]:
                books_ws.cell(row=r, column=ewb_col).value = m["ref"]
        st.download_button(
            "Download Books — with EWB No column",
            data=workbook_to_bytes(books_wb),
            file_name="Books - Matched.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with d2:
        ewb_wb = st.session_state["ewb_wb"]
        ewb_ws = ewb_wb.worksheets[
            [ws.title for ws in ewb_wb.worksheets].index(st.session_state["ewb_ws_title"])
        ]
        c_books = get_or_create_col(ewb_ws, "As per Books", "asperbooks")
        c_diff = get_or_create_col(ewb_ws, "Diff", "diff")
        c_voucher = get_or_create_col(ewb_ws, "Matched Voucher No", "matchedvoucherno")
        by_row = {m["r"]: m for m in res["matches"]}
        for r in range(2, ewb_ws.max_row + 1):
            m = by_row.get(r)
            ewb_ws.cell(row=r, column=c_books).value = m["sum"] if m else None
            ewb_ws.cell(row=r, column=c_diff).value = round(m["val"] - m["sum"], 2) if m else None
            ewb_ws.cell(row=r, column=c_voucher).value = m["vn"] if m else None
        st.download_button(
            "Download E-way Bills — with matches",
            data=workbook_to_bytes(ewb_wb),
            file_name="E way Bills - Matched.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

st.markdown(
    """
    <div style="margin-top:40px; padding-top:14px; border-top:1px dotted #B9B096; font-size:11.5px; color:#3A4A63;">
    Books rows are always summed by Voucher No. first. Whichever criteria you select in Step 3 then
    gate a match: <b>Amount</b> requires the voucher sum to be within your tolerance of the E-way Bill
    value; <b>Voucher No.</b> requires it to match the E-way Bill's Doc No.; <b>Party Name</b> requires
    a confirmed pairing (via the mapping file) between the Books and E-way Bill parties, so amounts are
    never compared across unrelated parties. You can select any one, two, or all three — amount
    closeness is always used to break ties either way. Each voucher is used at most once.
    </div>
    """,
    unsafe_allow_html=True,
)
