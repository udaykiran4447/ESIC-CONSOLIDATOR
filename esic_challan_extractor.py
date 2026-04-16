import streamlit as st
import pdfplumber
import pandas as pd
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="ESIC Challan Extractor", page_icon="📄", layout="centered")

st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'IBM Plex Sans', sans-serif;
    }
    .main { background-color: #f5f5f0; }
    .stApp { background-color: #f5f5f0; }
    
    h1 {
        font-family: 'IBM Plex Mono', monospace !important;
        font-size: 1.6rem !important;
        color: #1a1a2e !important;
        border-bottom: 3px solid #c0392b;
        padding-bottom: 0.4rem;
    }
    .stButton > button {
        background-color: #c0392b;
        color: white;
        border: none;
        border-radius: 4px;
        font-family: 'IBM Plex Mono', monospace;
        font-weight: 600;
        padding: 0.5rem 1.5rem;
        width: 100%;
    }
    .stButton > button:hover {
        background-color: #922b21;
    }
    .upload-box {
        background: white;
        border: 2px dashed #c0392b;
        border-radius: 8px;
        padding: 1rem;
        margin-bottom: 1rem;
    }
    .tag {
        display: inline-block;
        background: #1a1a2e;
        color: white;
        font-family: 'IBM Plex Mono', monospace;
        font-size: 0.75rem;
        padding: 2px 8px;
        border-radius: 3px;
        margin-right: 4px;
    }
    </style>
""", unsafe_allow_html=True)

st.title("ESIC Challan Extractor")
st.markdown("Upload one or more ESIC Challan PDFs to extract and export data to Excel.")

FIELDS = {
    "Employer's Code No": "Employer Code No",
    "Employer's Name": "Employer Name",
    "Challan Period": "Challan Period",
    "Challan Number": "Challan Number",
    "Challan Created Date": "Challan Created Date",
    "Challan Submitted Date": "Challan Submitted Date",
    "Amount Paid": "Amount Paid",
    "Transaction Number": "Transaction Number",
    "Transaction status": "Transaction Status",
}

def extract_from_pdf(file_bytes):
    data = {}
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    for field, col in FIELDS.items():
        pattern = re.escape(field) + r"[\s:]*([^\n]+)"
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            data[col] = match.group(1).strip().rstrip("*").strip()
        else:
            data[col] = ""
    return data

def create_excel(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "ESIC Challans"

    headers = list(FIELDS.values())

    # Styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1A1A2E")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Arial", size=10)
    alt_fill = PatternFill("solid", start_color="F2F2EF")
    center_align = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, record in enumerate(records, 2):
        fill = PatternFill("solid", start_color="FFFFFF") if row_idx % 2 == 0 else alt_fill
        for col_idx, header in enumerate(headers, 1):
            val = record.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = center_align
            cell.border = border
        ws.row_dimensions[row_idx].height = 20

    # Column widths
    col_widths = [22, 28, 14, 22, 22, 22, 14, 22, 26]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Total row
    total_row = len(records) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", start_color="C0392B")
    ws.cell(row=total_row, column=1).font = Font(name="Arial", bold=True, color="FFFFFF")
    ws.cell(row=total_row, column=1).alignment = center_align

    amt_col = headers.index("Amount Paid") + 1
    total_formula = f"=SUM({get_column_letter(amt_col)}2:{get_column_letter(amt_col)}{total_row-1})"
    total_cell = ws.cell(row=total_row, column=amt_col, value=total_formula)
    total_cell.font = Font(name="Arial", bold=True, color="FFFFFF")
    total_cell.fill = PatternFill("solid", start_color="C0392B")
    total_cell.alignment = center_align
    total_cell.border = border

    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Upload
uploaded_files = st.file_uploader(
    "Upload ESIC Challan PDFs",
    type=["pdf"],
    accept_multiple_files=True,
    help="You can upload multiple PDFs at once"
)

if uploaded_files:
    st.markdown(f"**{len(uploaded_files)} file(s) uploaded**")
    
    records = []
    errors = []

    for f in uploaded_files:
        try:
            record = extract_from_pdf(f.read())
            record["_filename"] = f.name
            records.append(record)
        except Exception as e:
            errors.append(f"{f.name}: {e}")

    if errors:
        for err in errors:
            st.error(f"⚠️ {err}")

    if records:
        display_records = [{k: v for k, v in r.items() if k != "_filename"} for r in records]
        df = pd.DataFrame(display_records)

        st.markdown("### Preview")
        st.dataframe(df, use_container_width=True)

        st.markdown(f"**{len(records)} record(s) extracted** | Total Amount: ₹{sum(float(r.get('Amount Paid', 0) or 0) for r in display_records):,.2f}")

        if st.button("⬇ Download Excel"):
            excel_file = create_excel(display_records)
            st.download_button(
                label="📥 Click to Save Excel File",
                data=excel_file,
                file_name="ESIC_Challans.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
else:
    st.info("Upload PDFs above to get started.")
